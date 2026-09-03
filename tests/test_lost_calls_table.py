import json
import stat
from datetime import datetime, timedelta, timezone
from urllib.parse import parse_qs, unquote, urlsplit

import pytest
from openpyxl import Workbook, load_workbook

from core import config
from core.lost_calls_table import (
    OUTPUT_HEADERS,
    TableFormatError,
    default_output_path,
    process_table,
)


def configure_dashboard(monkeypatch, tmp_path):
    config_path = tmp_path / "config.toml"
    config_path.write_text(
        """[services.zapis]
url = "https://grafana.test/d/calls?orgId=42&var-env=test"

[opensearch]
base_url = "https://opensearch.test/discover"

[opensearch.index_patterns]
sip_stack = "sip-stack-view"

[grafana.recording]
loki_datasource_uid = "loki-test"
""",
        encoding="utf-8",
    )
    monkeypatch.setattr(config, "CONFIG_PATH", config_path)


def create_source_xlsx(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Выгрузка"
    sheet.append(["Служебная строка"])
    sheet.append(
        [
            "Лишний столбец",
            "Номер пользователя",
            "Номер другой стороны",
            "Старт звонка (UTC)",
            "Продолжительность звонка",
            "Направление",
        ]
    )
    sheet.append(
        [
            "не переносить",
            79990000001,
            "+7 (999) 000-00-02",
            datetime(2026, 8, 1, 10, 0, 0),
            timedelta(seconds=95),
            "out",
        ]
    )
    sheet.append(
        [
            "не переносить",
            "79990000003",
            "79990000004",
            "2026-08-01T11:30:00Z",
            "00:02:10",
            "Входящий",
        ]
    )
    sheet.append(
        [
            "не переносить",
            "79990000005",
            "79990000006",
            "не дата",
            "15",
            "out",
        ]
    )
    workbook.save(path)
    workbook.close()


def link_params(cell):
    target = cell.hyperlink.target
    return parse_qs(urlsplit(target).query)


def test_process_xlsx_cleans_columns_and_builds_utc_links(
    monkeypatch,
    tmp_path,
):
    configure_dashboard(monkeypatch, tmp_path)
    source = tmp_path / "calls.xlsx"
    output = tmp_path / "cleaned.xlsx"
    create_source_xlsx(source)

    result = process_table(
        source,
        output,
        now=datetime(2026, 8, 2, 12, tzinfo=timezone.utc),
    )

    assert result.row_count == 3
    assert result.link_count == 6
    assert result.dropped_count == 0
    assert result.warnings == (
        "Строка 5: не заполнено или некорректно: старт звонка",
    )
    assert stat.S_IMODE(output.stat().st_mode) == 0o600

    workbook = load_workbook(output)
    sheet = workbook.active
    assert tuple(cell.value for cell in sheet[1]) == OUTPUT_HEADERS
    assert sheet.max_column == 8
    assert sheet.freeze_panes == "A2"
    assert sheet.sheet_view.showGridLines is False
    assert sheet["A2"].value == "79990000001"
    assert sheet["B2"].value == "+7 (999) 000-00-02"
    assert sheet["C2"].value == datetime(2026, 8, 1, 10, 0, 0)
    assert sheet["F2"].value == "Открыть"
    assert sheet["G2"].value == "Открыть"
    assert sheet["H2"].value == "Открыть"
    assert sheet["F4"].hyperlink is None
    assert sheet["G4"].hyperlink is None
    assert sheet["H4"].hyperlink is None

    outgoing = link_params(sheet["F2"])
    assert outgoing["timezone"] == ["UTC"]
    assert outgoing["var-phone"] == ["9990000001"]
    assert outgoing["var-second_phone"] == ["9990000002"]
    assert outgoing["from"] == ["2026-08-01T09:58:00.000Z"]
    assert outgoing["to"] == ["2026-08-01T11:30:00.000Z"]

    incoming = link_params(sheet["F3"])
    assert incoming["var-phone"] == ["9990000003"]
    assert incoming["var-second_phone"] == ["9990000004"]
    assert sheet["A2"].fill.fill_type is None
    assert sheet["B2"].fill.fill_type is None
    assert sheet["A3"].fill.fill_type is None
    assert sheet["B3"].fill.fill_type is None
    incoming_sip_link = unquote(sheet["G3"].hyperlink.target)
    incoming_mgw_link = sheet["H3"].hyperlink.target
    assert "sip-stack-view" in incoming_sip_link
    assert "*9990000003" in incoming_sip_link
    assert (
        "time:(from:'2026-08-01T14:28:00.000',to:'2026-08-01T16:00:00.000')"
        in incoming_sip_link
    )

    mgw_params = parse_qs(urlsplit(incoming_mgw_link).query)
    mgw_pane = json.loads(mgw_params["panes"][0])["A"]
    assert mgw_pane["range"] == {
        "from": "2026-08-01T11:28:00.000Z",
        "to": "2026-08-01T13:00:00.000Z",
    }
    assert mgw_pane["queries"][0]["expr"] == (
        '{unit="mgw.service"} |= "9990000003" |= "9990000004" | json'
    )
    workbook.close()


def test_csv_input_and_default_output_path(monkeypatch, tmp_path):
    configure_dashboard(monkeypatch, tmp_path)
    source = tmp_path / "calls.csv"
    source.write_text(
        "Номер пользователя;Номер другой стороны;Старт звонка;"
        "Продолжительность звонка;Направление звонка;Лишнее\n"
        "79990000001;79990000002;01.08.2026 10:00;30;out;удалить\n",
        encoding="utf-8",
    )

    result = process_table(
        source,
        now=datetime(2026, 8, 2, 12, tzinfo=timezone.utc),
    )

    assert result.output_path == default_output_path(source)
    assert result.row_count == 1
    assert result.link_count == 3
    assert result.dropped_count == 0
    workbook = load_workbook(result.output_path)
    assert workbook.active.max_column == 8
    assert workbook.active["F2"].hyperlink is not None
    assert workbook.active["G2"].hyperlink is not None
    assert workbook.active["H2"].hyperlink is not None
    workbook.close()


def test_unknown_direction_is_kept_as_metadata(monkeypatch, tmp_path):
    configure_dashboard(monkeypatch, tmp_path)
    source = tmp_path / "calls.csv"
    source.write_text(
        "Номер пользователя,Номер другой стороны,Старт звонка,"
        "Продолжительность звонка,Направление звонка\n"
        "79990000001,79990000002,2026-08-01 10:00,30,unknown\n",
        encoding="utf-8",
    )

    result = process_table(
        source,
        now=datetime(2026, 8, 2, 12, tzinfo=timezone.utc),
    )

    assert result.link_count == 3
    assert result.dropped_count == 0
    assert result.warnings == ()


def test_invalid_phone_lengths_do_not_generate_links(monkeypatch, tmp_path):
    configure_dashboard(monkeypatch, tmp_path)
    source = tmp_path / "calls.csv"
    source.write_text(
        "Номер пользователя,Номер другой стороны,Старт звонка,"
        "Продолжительность звонка,Направление звонка\n"
        "123,456,2026-08-01 10:00,30,out\n",
        encoding="utf-8",
    )

    result = process_table(
        source,
        now=datetime(2026, 8, 2, 12, tzinfo=timezone.utc),
    )

    assert result.link_count == 0
    assert result.warnings == (
        "Строка 2: не заполнено или некорректно: "
        "номер пользователя, номер другой стороны",
    )


def test_calls_older_than_five_days_are_removed(monkeypatch, tmp_path):
    configure_dashboard(monkeypatch, tmp_path)
    source = tmp_path / "calls.csv"
    source.write_text(
        "Номер пользователя,Номер другой стороны,Старт звонка,"
        "Продолжительность звонка,Направление звонка\n"
        "79990000010,79990000110,2026-08-23 11:59,30,out\n"
        "79990000011,79990000111,2026-08-23 12:00,30,out\n"
        "79990000012,79990000112,2026-08-27 12:00,30,out\n"
        "79990000013,79990000113,не дата,30,out\n",
        encoding="utf-8",
    )

    result = process_table(
        source,
        now=datetime(2026, 8, 28, 12, tzinfo=timezone.utc),
    )

    assert result.row_count == 3
    assert result.dropped_count == 1
    assert result.link_count == 6
    assert result.warnings == (
        "Строка 5: не заполнено или некорректно: старт звонка",
    )

    workbook = load_workbook(result.output_path)
    sheet = workbook.active
    assert [sheet.cell(row, 1).value for row in range(2, 5)] == [
        "79990000011",
        "79990000012",
        "79990000013",
    ]
    workbook.close()


def test_missing_required_column_is_rejected(tmp_path):
    source = tmp_path / "calls.csv"
    source.write_text(
        "Номер пользователя,Номер другой стороны,Старт звонка\n",
        encoding="utf-8",
    )

    with pytest.raises(TableFormatError, match="Не найдены обязательные столбцы"):
        process_table(source)


def test_source_file_cannot_be_overwritten(tmp_path):
    source = tmp_path / "calls.csv"
    source.write_text("", encoding="utf-8")

    with pytest.raises(TableFormatError, match="не должен перезаписывать"):
        process_table(source, source)
