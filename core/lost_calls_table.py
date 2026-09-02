import csv
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from core.config import (
    grafana_recording_loki_datasource_uid,
    service_url,
)
from core.time_windows import fmt_utc
from modules import find_call_in_logs, sip_stack_opensearch
from services.loki_explore import build_explore_url_from_dashboard


OUTPUT_HEADERS = (
    "Номер пользователя",
    "Номер другой стороны",
    "Старт звонка",
    "Продолжительность звонка",
    "Направление звонка",
    "Ссылка: Zapis",
    "Ссылка: SIP stack prod",
    "Ссылка: MGW",
)

LINK_SERVICES = (
    ("zapis", "Zapis"),
    ("sip_stack", "SIP stack prod"),
    ("mgw", "MGW"),
)

MAX_CALL_AGE = timedelta(days=5)
LINK_WINDOW_BEFORE = timedelta(minutes=2)
LINK_WINDOW_AFTER = timedelta(minutes=90)

SOURCE_FIELDS = (
    "user_phone",
    "other_phone",
    "call_start",
    "call_duration",
    "call_direction",
)

HEADER_ALIASES = {
    "user_phone": (
        "номер пользователя",
        "user number",
        "user_number",
    ),
    "other_phone": (
        "номер другой стороны",
        "номер второй стороны",
        "other party number",
        "other_number",
    ),
    "call_start": (
        "старт звонка",
        "старт звонка utc",
        "начало звонка",
        "call start",
        "call_start",
    ),
    "call_duration": (
        "продолжительность звонка",
        "длительность звонка",
        "call duration",
        "call_duration",
    ),
    "call_direction": (
        "направление звонка",
        "направление",
        "call direction",
        "call_direction",
    ),
}

@dataclass(frozen=True)
class SourceRow:
    source_row: int
    user_phone: object
    other_phone: object
    call_start: object
    call_duration: object
    call_direction: object


@dataclass(frozen=True)
class TableResult:
    output_path: Path
    row_count: int
    link_count: int
    dropped_count: int
    warnings: tuple[str, ...]


class TableFormatError(ValueError):
    pass


def normalize_header(value):
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").strip().lower().replace("ё", "е")
    text = re.sub(r"[()\[\]]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def resolve_headers(values):
    normalized_values = [normalize_header(value) for value in values]
    resolved = {}

    for field_name, aliases in HEADER_ALIASES.items():
        for column_index, value in enumerate(normalized_values):
            if value in aliases:
                resolved[field_name] = column_index
                break

    return resolved


def _missing_header_message(header_map):
    missing = [
        OUTPUT_HEADERS[index]
        for index, field_name in enumerate(SOURCE_FIELDS)
        if field_name not in header_map
    ]
    return "Не найдены обязательные столбцы: " + ", ".join(missing)


def _find_xlsx_header(sheet, max_rows=50):
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, max_rows)),
        start=1,
    ):
        header_map = resolve_headers([cell.value for cell in row])
        if len(header_map) == len(SOURCE_FIELDS):
            return row_number, header_map

    raise TableFormatError(_missing_header_message({}))


def _read_xlsx(path, sheet_name=None):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise TableFormatError(f"Лист не найден: {sheet_name}")
            sheets = [workbook[sheet_name]]
        else:
            sheets = list(workbook.worksheets)

        last_error = None
        for sheet in sheets:
            try:
                header_row, header_map = _find_xlsx_header(sheet)
            except TableFormatError as error:
                last_error = error
                continue

            rows = []
            for source_row, cells in enumerate(
                sheet.iter_rows(min_row=header_row + 1),
                start=header_row + 1,
            ):
                values = [
                    cells[header_map[field_name]].value
                    if header_map[field_name] < len(cells)
                    else None
                    for field_name in SOURCE_FIELDS
                ]
                if any(value not in (None, "") for value in values):
                    rows.append(SourceRow(source_row, *values))

            return sheet.title, rows

        if sheet_name and last_error:
            raise last_error
        raise TableFormatError(
            "Ни на одном листе не найдены все обязательные столбцы: "
            + ", ".join(OUTPUT_HEADERS[:5])
        )
    finally:
        workbook.close()


def _csv_dialect(sample, suffix):
    if suffix == ".tsv":
        return csv.excel_tab

    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        return csv.excel


def _read_csv(path):
    text = path.read_text(encoding="utf-8-sig")
    dialect = _csv_dialect(text[:4096], path.suffix.lower())
    all_rows = list(csv.reader(text.splitlines(), dialect=dialect))

    for header_index, values in enumerate(all_rows[:50]):
        header_map = resolve_headers(values)
        if len(header_map) != len(SOURCE_FIELDS):
            continue

        rows = []
        for source_row, values in enumerate(
            all_rows[header_index + 1 :],
            start=header_index + 2,
        ):
            selected = [
                values[header_map[field_name]]
                if header_map[field_name] < len(values)
                else None
                for field_name in SOURCE_FIELDS
            ]
            if any(value not in (None, "") for value in selected):
                rows.append(SourceRow(source_row, *selected))
        return path.stem, rows

    raise TableFormatError(_missing_header_message({}))


def read_source_rows(input_path, sheet_name=None):
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, sheet_name=sheet_name)
    if suffix in {".csv", ".tsv"}:
        if sheet_name:
            raise TableFormatError("Параметр --sheet применим только к XLSX")
        return _read_csv(path)

    raise TableFormatError("Поддерживаются файлы XLSX, XLSM, CSV и TSV")


def phone_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_phone(value):
    raw_value = phone_text(value)
    digits = re.sub(r"\D", "", raw_value)
    if len(digits) == 10:
        return "7" + digits
    if len(digits) == 11 and digits.startswith("8"):
        return "7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return digits
    return ""


def parse_utc_datetime(value):
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = str(value or "").strip()
        if not text:
            return None

        parsed = None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            pass

        if parsed is None:
            for date_format in (
                "%d.%m.%Y, %H:%M:%S.%f",
                "%d.%m.%Y, %H:%M:%S",
                "%d.%m.%Y, %H:%M",
                "%d.%m.%Y %H:%M:%S.%f",
                "%d.%m.%Y %H:%M:%S",
                "%d.%m.%Y %H:%M",
                "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M",
            ):
                try:
                    parsed = datetime.strptime(text, date_format)
                    break
                except ValueError:
                    continue

        if parsed is None:
            return None

    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def filter_recent_rows(rows, now=None):
    current = now or datetime.now(timezone.utc)
    if current.tzinfo is not None:
        current = current.astimezone(timezone.utc).replace(tzinfo=None)

    cutoff = current - MAX_CALL_AGE
    recent_rows = []
    dropped_count = 0
    for row in rows:
        call_start = parse_utc_datetime(row.call_start)
        if call_start is not None and call_start < cutoff:
            dropped_count += 1
            continue
        recent_rows.append(row)

    return recent_rows, dropped_count


def event_window(event_time):
    return event_time - LINK_WINDOW_BEFORE, event_time + LINK_WINDOW_AFTER


def format_opensearch_msk(value):
    msk_value = value.replace(tzinfo=timezone.utc).astimezone(
        ZoneInfo("Europe/Moscow")
    )
    return msk_value.strftime("%Y-%m-%dT%H:%M:%S.000")


def build_sip_stack_links(ctx, event_time):
    time_from, time_to = event_window(event_time)
    link = sip_stack_opensearch.build_one(
        ctx,
        format_opensearch_msk(time_from),
        format_opensearch_msk(time_to),
    )
    return [link] if link else []


def mgw_phone(value):
    phone = normalize_phone(value)
    if len(phone) == 11 and phone.startswith("7"):
        return phone[1:]
    return phone


def build_mgw_links(ctx, event_time):
    phone_a = mgw_phone(ctx.get("phone_a"))
    phone_b = mgw_phone(ctx.get("phone_b"))
    if not phone_a or not phone_b:
        return []

    time_from, time_to = event_window(event_time)
    return [
        build_explore_url_from_dashboard(
            service_url("zapis"),
            grafana_recording_loki_datasource_uid(),
            f'{{unit="mgw.service"}} |= "{phone_a}" |= "{phone_b}" | json',
            time_from=fmt_utc(time_from),
            time_to=fmt_utc(time_to),
        )
    ]


def build_links(row, window=None):
    user_phone = normalize_phone(row.user_phone)
    other_phone = normalize_phone(row.other_phone)
    event_time = parse_utc_datetime(row.call_start)

    missing = []
    if not user_phone:
        missing.append("номер пользователя")
    if not other_phone:
        missing.append("номер другой стороны")
    if event_time is None:
        missing.append("старт звонка")
    if missing:
        return {}, ("не заполнено или некорректно: " + ", ".join(missing),)

    ctx = {
        "msisdn": user_phone,
        "phone_a": user_phone,
        "phone_a_values": [user_phone],
        "phone_b": other_phone,
        "phone_b_values": [other_phone],
        "event_time": event_time,
        "event_datetimes": [event_time],
        "event_time_range": event_window(event_time),
        "tz": "UTC",
    }

    service_builders = {
        "zapis": lambda: find_call_in_logs.build(ctx),
        "sip_stack": lambda: build_sip_stack_links(ctx, event_time),
        "mgw": lambda: build_mgw_links(ctx, event_time),
    }

    links_by_service = {}
    warnings = []
    for service_name, service_title in LINK_SERVICES:
        try:
            links = service_builders[service_name]()
        except Exception as error:
            warnings.append(f"не удалось создать ссылку {service_title}: {error}")
            continue
        if links:
            links_by_service[service_name] = links[0]

    return links_by_service, tuple(warnings)


def _safe_sheet_title(value):
    title = re.sub(r"[\\/*?:\[\]]", " ", str(value)).strip()
    return (title or "Потерянные записи")[:31]


def _output_value(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def write_clean_workbook(
    rows,
    output_path,
    *,
    sheet_title,
    window=60,
    dropped_count=0,
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(sheet_title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    sheet.append(OUTPUT_HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(vertical="center", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    sheet.row_dimensions[1].height = 32

    thin_gray = Side(style="thin", color="D9E2F3")
    body_font = Font(name="Aptos", size=11, color="1F2937")
    warnings = []
    link_count = 0

    for output_row, source_row in enumerate(rows, start=2):
        links_by_service, row_warnings = build_links(source_row, window)
        values = (
            phone_text(source_row.user_phone),
            phone_text(source_row.other_phone),
            _output_value(parse_utc_datetime(source_row.call_start) or source_row.call_start),
            _output_value(source_row.call_duration),
            _output_value(source_row.call_direction),
            *(
                "Открыть" if service_name in links_by_service else ""
                for service_name, _service_title in LINK_SERVICES
            ),
        )
        sheet.append(values)

        for cell in sheet[output_row]:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(bottom=thin_gray)

        for column_index in (3, 4, 5):
            sheet.cell(output_row, column_index).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        sheet.cell(output_row, 1).number_format = "@"
        sheet.cell(output_row, 2).number_format = "@"
        if isinstance(sheet.cell(output_row, 3).value, (datetime, date)):
            sheet.cell(output_row, 3).number_format = "yyyy-mm-dd hh:mm:ss"
        if isinstance(source_row.call_duration, (datetime, time)):
            sheet.cell(output_row, 4).number_format = "hh:mm:ss"
        elif hasattr(source_row.call_duration, "total_seconds"):
            sheet.cell(output_row, 4).number_format = "[h]:mm:ss"

        for column_index, (service_name, _service_title) in enumerate(
            LINK_SERVICES,
            start=6,
        ):
            link = links_by_service.get(service_name)
            if not link:
                continue
            link_cell = sheet.cell(output_row, column_index)
            link_cell.hyperlink = link
            link_cell.font = Font(name="Aptos", size=11, color="0563C1", underline="single")
            link_count += 1
        warnings.extend(
            f"Строка {source_row.source_row}: {warning}"
            for warning in row_warnings
        )

    last_row = max(sheet.max_row, 1)
    last_column = len(OUTPUT_HEADERS)
    last_column_letter = get_column_letter(last_column)
    if rows:
        table = Table(
            displayName="LostCallsTable",
            ref=f"A1:{last_column_letter}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"

    for column_letter, width in zip(
        "ABCDEFGH",
        (22, 22, 24, 28, 23, 18, 20, 18),
    ):
        sheet.column_dimensions[column_letter].width = width

    file_descriptor = os.open(
        output_path,
        os.O_WRONLY | os.O_CREAT | os.O_TRUNC,
        0o600,
    )
    os.fchmod(file_descriptor, 0o600)
    with os.fdopen(file_descriptor, "wb") as file:
        workbook.save(file)
    workbook.close()
    return TableResult(
        output_path=output_path,
        row_count=len(rows),
        link_count=link_count,
        dropped_count=dropped_count,
        warnings=tuple(warnings),
    )


def default_output_path(input_path):
    path = Path(input_path)
    return path.with_name(f"{path.stem}.cleaned.xlsx")


def process_table(
    input_path,
    output_path=None,
    *,
    sheet_name=None,
    window=60,
    now=None,
):
    if window < 0:
        raise ValueError("Окно поиска не может быть отрицательным")

    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else default_output_path(input_path)
    if input_path.resolve() == output_path.resolve():
        raise TableFormatError("Выходной файл не должен перезаписывать исходную таблицу")

    source_title, rows = read_source_rows(input_path, sheet_name=sheet_name)
    rows, dropped_count = filter_recent_rows(rows, now=now)
    return write_clean_workbook(
        rows,
        output_path,
        sheet_title=f"{source_title} — очищено",
        window=window,
        dropped_count=dropped_count,
    )
